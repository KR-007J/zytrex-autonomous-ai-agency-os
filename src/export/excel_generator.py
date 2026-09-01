"""Automated multi-sheet Excel spreadsheet (.xlsx) generator with professional styling."""

from __future__ import annotations
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional, List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from src.database.models import Lead, OutreachDraft, ContactedMemory
from src.database.db import get_db_session

EXPORTS_DIR = Path(__file__).resolve().parent.parent.parent / "data" / "exports"


class ExcelReportGenerator:
    def __init__(self, exports_dir: Optional[Path] = None):
        self.exports_dir = exports_dir or EXPORTS_DIR
        self.exports_dir.mkdir(parents=True, exist_ok=True)

    def generate_daily_master_sheet(self, session: Session, filename: Optional[str] = None) -> Path:
        """Generate comprehensive, professionally styled .xlsx workbook."""
        wb = openpyxl.Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        # Style definitions
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        title_font = Font(name="Calibri", size=14, bold=True, color="0F172A")
        sub_font = Font(name="Calibri", size=9, italic=True, color="64748B")
        cell_font = Font(name="Calibri", size=10)
        bold_font = Font(name="Calibri", size=10, bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

        # ==============================================================================
        # SHEET 1: Global Business & Stores Intelligence
        # ==============================================================================
        ws_global = wb.create_sheet(title="Global Store & Tech Leads")
        ws_global.views.sheetView[0].showGridLines = True

        ws_global.merge_cells("A1:K1")
        ws_global["A1"] = f"Zytrex Global Business Intelligence & Store Registry — Generated {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws_global["A1"].font = title_font
        ws_global["A1"].alignment = left_align

        ws_global.merge_cells("A2:K2")
        ws_global["A2"] = "Multi-platform business directory covering OpenCart, WordPress/WooCommerce, Shopify, Supplements & Steroids Stores, and Global B2B SaaS."
        ws_global["A2"].font = sub_font
        ws_global["A2"].alignment = left_align

        headers_global = [
            "ID", "Store / Company Name", "Platform / CMS", "Category / Niche", "Region / Country",
            "Live Store URL", "Contact Email", "Contact Phone", "Detected Tech Stack", "Confidence", "Compliance Status"
        ]

        row_num = 4
        for col_idx, h in enumerate(headers_global, 1):
            cell = ws_global.cell(row=row_num, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        from src.database.models import GlobalEnterpriseLead
        global_leads = session.query(GlobalEnterpriseLead).order_by(GlobalEnterpriseLead.confidence_score.desc(), GlobalEnterpriseLead.created_at.desc()).all()
        for idx, g in enumerate(global_leads, 1):
            row_num += 1
            is_even = (idx % 2 == 0)

            tech_str = ""
            if g.tech_stack:
                try:
                    tech_str = ", ".join(json.loads(g.tech_stack))
                except Exception:
                    tech_str = str(g.tech_stack)

            row_values = [
                g.id,
                g.company_name,
                g.platform_cms,
                g.category,
                f"{g.region} ({g.country_code})",
                g.live_url,
                g.contact_email or "—",
                g.contact_phone or "—",
                tech_str or "Web Standard",
                f"{g.confidence_score}%",
                g.compliance_status,
            ]

            for col_idx, val in enumerate(row_values, 1):
                c = ws_global.cell(row=row_num, column=col_idx, value=val)
                c.font = cell_font
                c.border = thin_border
                if is_even:
                    c.fill = zebra_fill
                if col_idx in (1, 3, 10, 11):
                    c.alignment = center_align
                else:
                    c.alignment = left_align

        for col in ws_global.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_global.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # ==============================================================================
        # SHEET 2: Agency Leads Explorer
        # ==============================================================================
        ws_leads = wb.create_sheet(title="Agency Scout Pipeline")
        ws_leads.views.sheetView[0].showGridLines = True

        ws_leads.merge_cells("A1:J1")
        ws_leads["A1"] = f"Agency Real-Time Prospect Pipeline — Generated {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws_leads["A1"].font = title_font
        ws_leads["A1"].alignment = left_align

        headers_leads = [
            "ID", "Company Name", "Contact Person", "Email Address", "Phone Number",
            "Niche / Opportunity", "Website URL", "Lead Score", "Status", "Date Discovered"
        ]

        row_num = 3
        for col_idx, h in enumerate(headers_leads, 1):
            cell = ws_leads.cell(row=row_num, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        leads = session.query(Lead).order_by(Lead.lead_score.desc(), Lead.created_at.desc()).all()
        for idx, l in enumerate(leads, 1):
            row_num += 1
            is_even = (idx % 2 == 0)
            row_values = [
                l.id,
                l.company_name,
                l.contact_name or "—",
                l.email or "—",
                l.phone or "—",
                l.industry_tag or "Web Discovery",
                l.source_url or l.source_domain or "—",
                l.lead_score,
                l.status.upper(),
                l.created_at.strftime("%Y-%m-%d") if l.created_at else "—"
            ]

            for col_idx, val in enumerate(row_values, 1):
                c = ws_leads.cell(row=row_num, column=col_idx, value=val)
                c.font = cell_font
                c.border = thin_border
                if is_even:
                    c.fill = zebra_fill
                if col_idx in (1, 8, 9, 10):
                    c.alignment = center_align
                else:
                    c.alignment = left_align

                if col_idx == 8 and val >= 70:
                    c.font = Font(name="Calibri", size=10, bold=True, color="16A34A")
                elif col_idx == 9 and val == "CONTACTED":
                    c.font = Font(name="Calibri", size=10, bold=True, color="2563EB")

        for col in ws_leads.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_leads.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # ==============================================================================
        # SHEET 2: Cold Email Outreach & Campaign Log
        # ==============================================================================
        ws_email = wb.create_sheet(title="Outreach Campaign Log")
        ws_email.views.sheetView[0].showGridLines = True

        ws_email.merge_cells("A1:G1")
        ws_email["A1"] = "Cold Email & Manual Outreach Execution Log"
        ws_email["A1"].font = title_font

        headers_email = ["Draft ID", "Lead ID", "Target Company", "Email Subject", "Outreach Status", "Created Date", "Sent Date"]
        row_num = 3
        for col_idx, h in enumerate(headers_email, 1):
            cell = ws_email.cell(row=row_num, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        drafts = session.query(OutreachDraft).order_by(OutreachDraft.created_at.desc()).all()
        for idx, d in enumerate(drafts, 1):
            row_num += 1
            row_values = [
                d.id,
                d.lead_id,
                d.lead.company_name if d.lead else "—",
                d.subject,
                d.status.upper(),
                d.created_at.strftime("%Y-%m-%d %H:%M") if d.created_at else "—",
                d.sent_at.strftime("%Y-%m-%d %H:%M") if d.sent_at else "—"
            ]
            for col_idx, val in enumerate(row_values, 1):
                c = ws_email.cell(row=row_num, column=col_idx, value=val)
                c.font = cell_font
                c.border = thin_border
                c.alignment = center_align if col_idx in (1, 2, 5, 6, 7) else left_align

        for col in ws_email.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_email.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # ==============================================================================
        # SHEET 3: Platform Analytics & Breakdown
        # ==============================================================================
        ws_stats = wb.create_sheet(title="Platform Analytics")
        ws_stats.views.sheetView[0].showGridLines = True

        ws_stats["A1"] = "Social Platform Lead Distribution"
        ws_stats["A1"].font = title_font

        ws_stats["A3"] = "Metric"
        ws_stats["B3"] = "Value"
        ws_stats["A3"].font = header_font
        ws_stats["A3"].fill = header_fill
        ws_stats["B3"].font = header_font
        ws_stats["B3"].fill = header_fill

        total_leads = len(leads)
        with_email = sum(1 for l in leads if l.email)
        with_phone = sum(1 for l in leads if l.phone)
        contacted = sum(1 for l in leads if l.status == "contacted")

        stats_rows = [
            ("Total Scraped Leads", total_leads),
            ("Verified Email Addresses", with_email),
            ("Direct Phone Numbers", with_phone),
            ("Total Leads Contacted", contacted),
            ("Pending Outreach Drafts", len(drafts)),
            ("Email Extraction Rate", f"{(with_email / total_leads * 100):.1f}%" if total_leads else "0%"),
        ]

        for r_idx, (k, v) in enumerate(stats_rows, 4):
            c1 = ws_stats.cell(row=r_idx, column=1, value=k)
            c2 = ws_stats.cell(row=r_idx, column=2, value=v)
            c1.font = bold_font
            c2.font = cell_font
            c1.border = thin_border
            c2.border = thin_border
            c2.alignment = center_align

        ws_stats.column_dimensions["A"].width = 30
        ws_stats.column_dimensions["B"].width = 20

        # Save files
        now_str = datetime.now().strftime("%Y-%m-%d")
        daily_filename = filename or f"Leads_Master_{now_str}.xlsx"
        daily_path = self.exports_dir / daily_filename
        latest_path = self.exports_dir / "leads_latest.xlsx"

        wb.save(daily_path)
        wb.save(latest_path)

        return daily_path
