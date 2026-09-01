"""Comprehensive Test Suite for Enterprise Global Lead Gen Platform."""

import pytest
from src.scraper.nlp_classifier import BusinessCategoryClassifier, RegionResolver
from src.compliance.guardrails import ComplianceManager
from src.database.db import init_db, get_db_session, GlobalLeadRepository, SuppressionRepository, IngestionJobRepository


@pytest.fixture(autouse=True)
def setup_db():
    init_db()


def test_category_classifier_accuracy():
    test_cases = [
        ("Online boutique for luxury dresses, fashion bags, and jewelry cart checkout", "dresses.com", "Retail & E-Commerce"),
        ("Dental clinic providing teeth cleaning, orthodontics, and implants by certified dentists", "mumbaidental.in", "Healthcare & Medical"),
        ("B2B SaaS cloud platform with REST APIs, devops analytics, and developer tools", "cloudscale.io", "B2B SaaS & Tech"),
        ("Commercial law firm specializing in corporate litigation, tax advisory, and compliance", "legalpartners.co.uk", "Legal & Professional Services"),
        ("Global freight shipping, container logistics, freight forwarding, and warehousing", "freightcargo.com", "Logistics & Supply Chain"),
    ]

    correct = 0
    for text, domain, expected in test_cases:
        predicted, conf = BusinessCategoryClassifier.classify(text, domain)
        if predicted == expected:
            correct += 1

    accuracy = correct / len(test_cases)
    assert accuracy >= 0.80, f"Category classifier accuracy {accuracy*100}% below 80% threshold."


def test_region_resolver_rules():
    assert RegionResolver.resolve(phone="+49 30 123456")["country_code"] == "DE"
    assert RegionResolver.resolve(phone="+91 22 98765432")["country_code"] == "IN"
    assert RegionResolver.resolve(domain="sample.co.uk")["country_code"] == "UK"
    assert RegionResolver.resolve(domain="shop.de")["country_code"] == "DE"
    assert RegionResolver.resolve(address_snippet="Headquarters in Berlin, Germany")["country_code"] == "DE"
    assert RegionResolver.resolve(address_snippet="Located in Bangalore, Karnataka")["country_code"] == "IN"


def test_suppression_and_erasure():
    with get_db_session() as session:
        # Add a lead
        lead, _ = GlobalLeadRepository.upsert_lead(session, {
            "company_name": "Test Opt Out Corp",
            "source_domain": "testoptout.com",
            "live_url": "https://testoptout.com",
            "category": "Retail & E-Commerce",
            "region": "Europe (EU)",
            "country_code": "DE",
            "contact_email": "info@testoptout.com",
            "confidence_score": 90,
        })

        # Check before suppression
        leads, count_before = GlobalLeadRepository.get_leads(session, search="testoptout.com")
        assert count_before == 1

        # Add to suppression
        SuppressionRepository.add_suppression(session, "testoptout.com", reason="GDPR Erasure Request")

        # Check after suppression (should be excluded from active search)
        leads_after, count_after = GlobalLeadRepository.get_leads(session, search="testoptout.com")
        assert count_after == 0

        # Remove suppression
        SuppressionRepository.remove_suppression(session, "testoptout.com")
        leads_restored, count_restored = GlobalLeadRepository.get_leads(session, search="testoptout.com")
        assert count_restored == 1


def test_compliance_audit_guardrails():
    audit_eu = ComplianceManager.audit_record({
        "contact_email": "contact@zalando.de",
        "country_code": "DE",
    })
    assert audit_eu["is_eu_resident"] is True
    assert audit_eu["is_corporate_generic"] is True
    assert "Article 6(1)(f) GDPR" in audit_eu["legal_basis"]

    audit_us = ComplianceManager.audit_record({
        "contact_email": "sales@stripe.com",
        "country_code": "US",
    })
    assert audit_us["is_eu_resident"] is False
    assert audit_us["is_corporate_generic"] is True


def test_platform_and_niche_filtering():
    with get_db_session() as session:
        # Test OpenCart platform filtering
        opencart_leads, opencart_total = GlobalLeadRepository.get_leads(session, platform_cms="OpenCart")
        assert opencart_total >= 10, f"Expected >= 10 OpenCart leads, found {opencart_total}"
        for lead in opencart_leads:
            assert lead.platform_cms == "OpenCart"

        # Test WordPress platform filtering
        wp_leads, wp_total = GlobalLeadRepository.get_leads(session, platform_cms="WordPress")
        assert wp_total >= 10, f"Expected >= 10 WordPress leads, found {wp_total}"
        for lead in wp_leads:
            assert lead.platform_cms == "WordPress"

        # Test Steroids & Fitness Supplements niche filtering
        supp_leads, supp_total = GlobalLeadRepository.get_leads(session, category="Steroids & Fitness Supplements")
        assert supp_total >= 10, f"Expected >= 10 Supplements leads, found {supp_total}"
        for lead in supp_leads:
            assert lead.category == "Steroids & Fitness Supplements"


def test_excel_export_multi_sheet():
    from src.export.excel_generator import ExcelReportGenerator
    import openpyxl

    with get_db_session() as session:
        gen = ExcelReportGenerator()
        path = gen.generate_daily_master_sheet(session)
        assert path.exists()

        wb = openpyxl.load_workbook(path)
        sheet_names = wb.sheetnames
        assert "Global Store & Tech Leads" in sheet_names
        assert "Agency Scout Pipeline" in sheet_names
        assert "Outreach Campaign Log" in sheet_names

        ws = wb["Global Store & Tech Leads"]
        assert ws["B4"].value == "Store / Company Name"
        assert ws["C4"].value == "Platform / CMS"
        assert ws["D4"].value == "Category / Niche"

