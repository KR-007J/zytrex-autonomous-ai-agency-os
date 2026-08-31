"""Unit tests for Excel spreadsheet (.xlsx) generation."""

import pytest
import openpyxl
from pathlib import Path
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from src.database.models import Base, Lead, OutreachDraft
from src.export.excel_generator import ExcelReportGenerator


@pytest.fixture
def db_session():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine)
    session = SessionLocal()
    try:
        yield session
    finally:
        session.close()


def test_excel_generation(db_session, tmp_path):
    # Insert sample leads
    lead1 = Lead(
        company_name="Acme AI",
        contact_name="Alice Smith",
        email="alice@acme.ai",
        phone="+1-555-0101",
        industry_tag="LinkedIn",
        lead_score=95,
        status="new",
    )
    lead2 = Lead(
        company_name="CloudCore",
        contact_name="Bob Jones",
        email="bob@cloudcore.io",
        industry_tag="Twitter",
        lead_score=80,
        status="contacted",
    )
    db_session.add_all([lead1, lead2])
    db_session.commit()

    generator = ExcelReportGenerator(exports_dir=tmp_path)
    excel_path = generator.generate_daily_master_sheet(db_session, filename="test_leads.xlsx")

    assert excel_path.exists()
    assert excel_path.suffix == ".xlsx"

    # Verify openpyxl can read the sheets
    wb = openpyxl.load_workbook(excel_path)
    assert "All Verified Leads" in wb.sheetnames
    assert "Outreach Campaign Log" in wb.sheetnames
    assert "Platform Analytics" in wb.sheetnames

    ws_leads = wb["All Verified Leads"]
    assert ws_leads.max_row >= 5  # Title + headers + 2 leads
